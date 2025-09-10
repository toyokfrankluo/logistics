from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import math

# =========================
# Excel 导出
# =========================
def export_invoice(company_info, customer_name, orders, bank_accounts,
                   filename="账单.xlsx", date_from=None, date_to=None):
    """
    导出账单 Excel（xlsx）。
    company_info: dict 或 str。如果是 dict，尝试读取 key "公司名称" 和 "地址"；
                  如果是字符串则直接当作公司名。
    customer_name: 客户名称字符串
    orders: 列表，每项为 dict，包含 keys: date, order_no, service_no, qty, type,
           weight, destination, channel, fee (float), summary
    bank_accounts: BankAccount ORM 对象列表（或元素含 bank_name, account_name, account_no 属性/键）
    filename: 保存文件名（返回值也是此路径）
    date_from/date_to: 账单导出的日期范围
    """

    # 解析 company_info（兼容 dict 或 str）
    if isinstance(company_info, dict):
        company_name = company_info.get("公司名称") or company_info.get("name") or "公司名称"
        company_addr = company_info.get("地址") or company_info.get("address") or ""
    else:
        company_name = str(company_info or "")
        company_addr = ""

    wb = Workbook()
    ws = wb.active
    ws.title = "账单"

    # 样式
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    header_font = Font(size=12, bold=True)
    thin = Side(border_style="thin", color="000000")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 固定列数
    MAX_COL = 11

    # ===== 顶部公司名 =====
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=MAX_COL)
    c = ws.cell(row=1, column=1, value=company_name)
    c.alignment = align_center
    c.font = Font(size=18, bold=True)

    # ===== 标题（第 2 行）=====
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=MAX_COL)
    sub = ws.cell(row=2, column=1, value="快件运费收款通知单")
    sub.alignment = align_center
    sub.font = Font(size=14, bold=True)

    # ===== 客户名 + 日期范围（第 3 行）=====
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=MAX_COL - 1)
    client_cell = ws.cell(row=3, column=1, value=f"客户名称: {customer_name}")
    client_cell.alignment = align_left
    client_cell.font = Font(size=11)

    # 日期范围
    if date_from or date_to:
        if isinstance(date_from, datetime):
            date_from_str = date_from.strftime("%Y-%m-%d")
        else:
            date_from_str = str(date_from) if date_from else ""
        if isinstance(date_to, datetime):
            date_to_str = date_to.strftime("%Y-%m-%d")
        else:
            date_to_str = str(date_to) if date_to else ""
        date_range_str = f"{date_from_str} ~ {date_to_str}" if (date_from_str or date_to_str) else ""
    else:
        date_range_str = datetime.today().strftime("%Y-%m-%d")

    ws.cell(row=3, column=MAX_COL, value=f"账单日期: {date_range_str}").alignment = align_right

    # ===== 表头 =====
    headers = ["序号", "日期", "订单号", "服务商单号", "件数", "类型",
               "计费重/Kg", "目的地", "运输渠道", "合计费用", "账单摘要"]
    header_row = 4
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.alignment = align_center
        cell.font = header_font
        cell.border = thin_border

    # ===== 写订单数据 =====
    start_data_row = header_row + 1
    total_fee = 0.0
    total_weight = 0.0
    for idx, order in enumerate(orders, start=1):
        r = start_data_row + (idx - 1)
        ws.cell(row=r, column=1, value=idx).alignment = align_center
        ws.cell(row=r, column=2, value=order.get("date", "")).alignment = align_center
        ws.cell(row=r, column=3, value=order.get("order_no", "")).alignment = align_left
        ws.cell(row=r, column=4, value=order.get("service_no", "")).alignment = align_left
        ws.cell(row=r, column=5, value=order.get("qty", "")).alignment = align_center
        ws.cell(row=r, column=6, value=order.get("type", "")).alignment = align_center
        w = order.get("weight", 0) or 0
        ws.cell(row=r, column=7, value=w).alignment = align_center
        ws.cell(row=r, column=8, value=order.get("destination", "")).alignment = align_center
        ws.cell(row=r, column=9, value=order.get("channel", "")).alignment = align_center
        fee_val = float(order.get("fee") or 0.0)
        fee_cell = ws.cell(row=r, column=10, value=fee_val)
        fee_cell.number_format = u'￥#,##0.00'
        fee_cell.alignment = align_right
        ws.cell(row=r, column=11, value=order.get("summary", "")).alignment = align_left

        for col_idx in range(1, MAX_COL + 1):
            ws.cell(row=r, column=col_idx).border = thin_border

        total_fee += fee_val
        total_weight += float(w)

    after_data_row = start_data_row + max(0, len(orders)) + 1

    # ===== 合计 =====
    ws.cell(row=after_data_row, column=1, value="总票数:").alignment = align_left
    ws.cell(row=after_data_row, column=2, value=len(orders)).alignment = align_left

    ws.cell(row=after_data_row + 1, column=1, value="总重量:").alignment = align_left
    ws.cell(row=after_data_row + 1, column=2, value=f"{total_weight:.3f} Kg").alignment = align_left

    ws.cell(row=after_data_row + 2, column=1, value="总费用:").alignment = align_left
    total_fee_cell = ws.cell(row=after_data_row + 2, column=2, value=total_fee)
    total_fee_cell.number_format = u'￥#,##0.00'
    total_fee_cell.alignment = align_right

    # ===== 银行账户 =====
    bank_start_row = after_data_row + 4
    ws.cell(row=bank_start_row, column=1, value="收款银行账号").font = Font(bold=True)
    cur = bank_start_row + 1
    for ba in bank_accounts:
        if hasattr(ba, "bank_name"):
            bank_name = getattr(ba, "bank_name")
            account_name = getattr(ba, "account_name")
            account_no = getattr(ba, "account_no")
        elif isinstance(ba, dict):
            bank_name = ba.get("bank_name", "")
            account_name = ba.get("account_name", "")
            account_no = ba.get("account_no", "")
        else:
            bank_name = str(getattr(ba, "bank_name", "") or "")
            account_name = str(getattr(ba, "account_name", "") or "")
            account_no = str(getattr(ba, "account_no", ""))

        ws.cell(row=cur, column=1, value=f"开户行：{bank_name}").alignment = align_left
        ws.cell(row=cur + 1, column=1, value=f"户名：{account_name}").alignment = align_left
        ws.cell(row=cur + 2, column=1, value=f"账号：{account_no}").alignment = align_left
        cur += 4

    # ===== 页脚 =====
    footer_row = max(cur, bank_start_row + 1) + 2
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=MAX_COL)
    footer_cell = ws.cell(row=footer_row, column=1,
                          value=f"公司地址：{company_addr or '深圳市宝安区福永镇福海街道同富路3号惠明盛工业园5栋一楼'}")
    footer_cell.alignment = Alignment(horizontal="center", vertical="center")
    footer_cell.font = Font(size=9, color="808080")

    # ===== 列宽 =====
    col_widths = {
        1: 5,    # 序号
        2: 12,   # 日期
        3: 18,   # 订单号
        4: 18,   # 服务商单号
        5: 8,    # 件数
        6: 10,   # 类型
        7: 12,   # 计费重
        8: 12,   # 目的地
        9: 14,   # 运输渠道
        10: 14,  # 合计费用
        11: 40   # 摘要（加宽）
    }
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(filename)
    return filename


# =========================
# 轨迹查询相关
# =========================
from flask import render_template

def call_gettrack(carrier_id, tracking_number, agent_id=None):
    """
    轨迹查询统一接口（占位实现）。
    这里可以接第三方物流 API，也可以返回本地假数据。
    """
    # 你以后可以改成调用真实 API
    return {
        "tracking_number": tracking_number,
        "tracks": [
            {"time": "2025-09-10 10:00", "status": "包裹已揽收"},
            {"time": "2025-09-11 15:00", "status": "运输中"},
            {"time": "2025-09-12 09:30", "status": "派送中"},
        ]
    }

def format_tracks_from_data(data):
    """
    把 call_gettrack 返回的结果转成前端友好的格式。
    """
    if not data or "tracks" not in data:
        return []
    tracks = data["tracks"]
    return "\n".join([f"{t['time']} - {t['status']}" for t in tracks])

def render_template_safe(template_name, **context):
    """
    渲染模板的安全封装，避免出错时崩溃。
    """
    try:
        return render_template(template_name, **context)
    except Exception as e:
        return f"模板渲染失败: {e}"