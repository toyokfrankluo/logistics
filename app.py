# app.py â€” å®Œæ•´åç«¯ï¼ˆå« 5 å¤§éœ€æ±‚ï¼‰
import os
import io
import json
import time
import sqlite3
import traceback
from datetime import datetime
from pathlib import Path
from contextlib import closing
from flask_migrate import Migrate

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from werkzeug.security import generate_password_hash, check_password_hash

from flask import (
    Flask, request, render_template, render_template_string,
    redirect, url_for, flash, send_file, abort
)
from flask_cors import CORS


try:
    import pandas as pd
except Exception:
    pd = None

from flask_sqlalchemy import SQLAlchemy
from flask_admin import Admin
from flask_admin.contrib.sqla import ModelView
from flask_login import (
    LoginManager, UserMixin, login_user, logout_user, login_required, current_user
)
from jinja2 import TemplateNotFound

# ä½¿ç”¨ä½ æä¾›çš„ models.py å®šä¹‰
from models import db, Customer, CarrierAgent, Shipment, ManualTrack, BankAccount

# ------------------------------
# åŸºæœ¬åº”ç”¨ä¸é…ç½®
# ------------------------------
load_dotenv()

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
TEMPLATE_DIR = os.path.join(BASE_DIR, "templates")
STATIC_DIR = os.path.join(BASE_DIR, "static")

app = Flask(__name__, template_folder=TEMPLATE_DIR, static_folder=STATIC_DIR)
app.secret_key = os.getenv("FLASK_SECRET", "change-this-secret-for-dev")

CORS(app, resources={r"/*": {"origins": "*"}}, supports_credentials=True)

from views import views
app.register_blueprint(views)

app.config["JSON_AS_ASCII"] = False
app.config["TEMPLATES_AUTO_RELOAD"] = True
app.config["SQLALCHEMY_DATABASE_URI"] = os.getenv("DATABASE_URL", "sqlite:///logistics.db")
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db.init_app(app)
migrate = Migrate(app, db)

# ç™»å½•
login_manager = LoginManager(app)
login_manager.login_view = "login"

# ç®¡ç†åå°
admin = Admin(app, name="åå°ç®¡ç†", template_mode="bootstrap4")

# ç¯å¢ƒé…ç½®
DEFAULT_API_URL = os.getenv("API_URL", "http://ywsl.rtb56.com/webservice/PublicService.asmx/ServiceInterfaceUTF8")
CARRIERS_LIST = [c.strip() for c in os.getenv("CARRIERS", "rtb56").split(",") if c.strip()]
CARRIERS = {}
for cid in CARRIERS_LIST:
    key_prefix = cid.upper()
    CARRIERS[cid] = {
        "id": cid,
        "name": os.getenv(f"{key_prefix}_NAME", cid),
        "token": os.getenv(f"{key_prefix}_TOKEN"),
        "key": os.getenv(f"{key_prefix}_KEY"),
        "api_url": os.getenv(f"{key_prefix}_API_URL", DEFAULT_API_URL)
    }

PUBLIC_MODE = os.getenv("PUBLIC_MODE", "0") == "1"  # å…¬å…±æŸ¥è¯¢é¡µå±è”½æ‰¹é‡

# ç®€å•ç¼“å­˜
CACHE = {}
CACHE_TTL = int(os.getenv("CACHE_TTL", "600"))

# 17Track APIé…ç½®
TRACK17_API_KEY = os.getenv("TRACK17_API_KEY", "")
TRACK17_API_URL = os.getenv("TRACK17_API_URL", "https://api.17track.net/track/v2/gettrackinfo")

# ------------------------------
# å·¥å…·ï¼šsqlite æ–‡ä»¶è·¯å¾„
# ------------------------------
def get_sqlite_path_from_uri(uri: str):
    if not uri:
        return None
    uri = uri.strip()
    if uri.startswith("sqlite:///"):
        return uri.replace("sqlite:///", "", 1)
    if uri.startswith("sqlite:////"):
        return "/" + uri.replace("sqlite:////", "", 1)
    return None

# ------------------------------
# ç”¨æˆ·æ¨¡å‹ï¼ˆä»…åç«¯ä½¿ç”¨ï¼‰
# ------------------------------
class User(UserMixin, db.Model):
    __tablename__ = "users"
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    is_admin = db.Column(db.Boolean, default=False)

    def set_password(self, raw):
        self.password_hash = generate_password_hash(raw)

    def check_password(self, raw):
        return check_password_hash(self.password_hash, raw)

# Admin
admin.add_view(ModelView(User, db.session))
admin.add_view(ModelView(Customer, db.session))
admin.add_view(ModelView(CarrierAgent, db.session))
admin.add_view(ModelView(Shipment, db.session))
admin.add_view(ModelView(ManualTrack, db.session))
admin.add_view(ModelView(BankAccount, db.session))

# ------------------------------
# ç™»å½•å›è°ƒ
# ------------------------------
@login_manager.user_loader
def load_user(user_id):
    try:
        return User.query.get(int(user_id))
    except Exception:
        return None

# ------------------------------
# æ¨¡æ¿å®‰å…¨æ¸²æŸ“
# ------------------------------
def render_template_safe(template_name, **context):
    try:
        return render_template(template_name, **context)
    except TemplateNotFound:
        fallback = f"""
        <!doctype html>
        <html>
        <head><meta charset="utf-8"><title>{template_name} - å ä½</title></head>
        <body style="font-family: -apple-system, Arial;">
        <h2>ç¼ºå°‘é¡µé¢æ¨¡æ¿ï¼š{template_name}</h2>
        <p>è¯·åœ¨ <code>templates/</code> ç›®å½•ä¸­æ·»åŠ  <strong>{template_name}</strong> æ¨¡æ¿æ–‡ä»¶ã€‚</p>
        <pre>æ¸²æŸ“æ•°æ®ï¼ˆè°ƒè¯•ç”¨ï¼‰:</pre>
        <div style="white-space:pre-wrap;border:1px solid #ddd;padding:10px;">{context}</div>
        <p><a href="{url_for('index')}">è¿”å›é¦–é¡µ</a></p>
        </body>
        </html>
        """
        return render_template_string(fallback)

# ------------------------------
# å¯åŠ¨æ—¶è‡ªåŠ¨è¡¥åˆ—ï¼ˆSQLiteï¼‰
# ------------------------------
def ensure_sqlite_columns():
    uri = app.config.get("SQLALCHEMY_DATABASE_URI", "")
    sqlite_path = get_sqlite_path_from_uri(uri)
    if not sqlite_path:
        app.logger.info("é SQLiteï¼Œè·³è¿‡è‡ªåŠ¨è¡¥åˆ—ã€‚")
        return

    db_file = Path(sqlite_path)
    if not db_file.exists():
        app.logger.info(f"SQLite {sqlite_path} ä¸å­˜åœ¨ï¼Œcreate_all å·²åˆ›å»ºã€‚")
        return

    app.logger.info(f"å¼€å§‹æ£€æŸ¥å¹¶è¡¥åˆ— SQLite æ•°æ®åº“: {sqlite_path}")
    try:
        with closing(sqlite3.connect(str(db_file))) as conn:
            cur = conn.cursor()

            def table_has_col(table, col):
                cur.execute(f"PRAGMA table_info('{table}')")
                rows = cur.fetchall()
                cols = [r[1] for r in rows]
                return col in cols

            patches = {
                "carrier_agent": [
                    ("api_url", "TEXT"), ("username", "TEXT"), ("password", "TEXT"),
                    ("app_key", "TEXT"), ("app_token", "TEXT"), ("customer_code", "TEXT"),
                    ("supports_api", "INTEGER DEFAULT 1"), ("is_active", "INTEGER DEFAULT 1")
                ],
                "shipment": [
                    ("customer_id", "INTEGER"), ("agent_id", "INTEGER"), ("carrier_id", "TEXT"),
                    ("origin", "TEXT"), ("destination", "TEXT"), ("channel", "TEXT"),
                    ("product_type", "TEXT"), ("pieces", "INTEGER DEFAULT 1"),
                    ("weight", "REAL DEFAULT 0"), ("unit_price", "REAL DEFAULT 0"),
                    ("fee", "REAL DEFAULT 0"), ("surcharge_extra", "REAL DEFAULT 0"),
                    ("operation_fee", "REAL DEFAULT 0"), ("high_value_fee", "REAL DEFAULT 0"),
                    ("status", "TEXT"), ("note", "TEXT"),
                    ("agent_tracking_number", "TEXT"), ("third_party_tracking_number", "TEXT"),
                    ("created_at", "TEXT"), ("updated_at", "TEXT")
                ],
                "manual_track": [
                    ("shipment_id", "INTEGER"), ("happen_time", "TEXT"),
                    ("location", "TEXT"), ("description", "TEXT"), ("author", "TEXT"),
                    ("created_at", "TEXT"), ("updated_at", "TEXT")
                ],
                "bank_account": [
                    ("account_type", "TEXT"), ("bank_name", "TEXT"),
                    ("account_name", "TEXT"), ("account_no", "TEXT"),
                    ("is_default", "INTEGER DEFAULT 0"), ("remark", "TEXT"),
                    ("created_at", "TEXT"), ("updated_at", "TEXT")
                ],
                "customer": [
                    ("bank_info", "TEXT"),
                    ("created_at", "TEXT"), ("updated_at", "TEXT")
                ]
            }

            for table, cols in patches.items():
                cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table,))
                if cur.fetchone() is None:
                    app.logger.info(f"è¡¨ {table} ä¸å­˜åœ¨ï¼Œè·³è¿‡è¡¥åˆ—ï¼ˆcreate_all å·²å¤„ç†ï¼‰ã€‚")
                    continue
                app.logger.info(f"æ£€æŸ¥è¡¨ {table} çš„åˆ—...")
                for col, typ in cols:
                    if not table_has_col(table, col):
                        sql = f"ALTER TABLE {table} ADD COLUMN {col} {typ}"
                        try:
                            cur.execute(sql)
                            app.logger.info(f"æˆåŠŸä¸º {table} æ·»åŠ åˆ— {col} ({typ})")
                        except Exception as e:
                            app.logger.error(f"ä¸º {table} æ·»åŠ åˆ— {col} å¤±è´¥: {str(e)}")
                    else:
                        app.logger.debug(f"è¡¨ {table} å·²å­˜åœ¨åˆ— {col}ï¼Œè·³è¿‡ã€‚")

            conn.commit()
            app.logger.info("è¡¥åˆ—æ“ä½œå®Œæˆã€‚")
    except Exception as e:
        app.logger.exception(f"è¡¥åˆ—è¿‡ç¨‹ä¸­å‘ç”Ÿå…¨å±€é”™è¯¯: {str(e)}")

# ------------------------------
# 17Track API è°ƒç”¨å‡½æ•°
# ------------------------------
def call_17track(tracking_number):
    """è°ƒç”¨17Track APIè·å–è½¨è¿¹ä¿¡æ¯"""
    if not TRACK17_API_KEY:
        return {"error": "17Track APIå¯†é’¥æœªé…ç½®"}
    
    try:
        headers = {
            "Content-Type": "application/json",
            "17token": TRACK17_API_KEY
        }
        
        payload = {
            "number": tracking_number,
            "carrier": None  # è‡ªåŠ¨è¯†åˆ«å¿«é€’å…¬å¸
        }
        
        response = requests.post(TRACK17_API_URL, headers=headers, json=payload, timeout=15)
        
        if response.status_code != 200:
            return {"error": f"17Track APIè¿”å›é”™è¯¯ {response.status_code}"}
            
        data = response.json()
        
        # æ£€æŸ¥APIå“åº”çŠ¶æ€
        if data.get("status") != 200:
            return {"error": f"17Track APIé”™è¯¯: {data.get('message', 'æœªçŸ¥é”™è¯¯')}"}
            
        # è§£æè½¨è¿¹æ•°æ®
        tracks = []
        if data.get("data") and isinstance(data["data"], list) and len(data["data"]) > 0:
            for event in data["data"][0].get("track", []):
                tracks.append({
                    "track_occur_date": event.get("time", ""),
                    "track_location": event.get("location", ""),
                    "track_description": event.get("description", event.get("info", ""))
                })
        
        return {
            "success": "1",
            "cnmessage": "17Trackè½¨è¿¹æŸ¥è¯¢æˆåŠŸ",
            "data": [{"details": tracks}]
        }
        
    except Exception as e:
        return {"error": f"17Track APIè¯·æ±‚å¤±è´¥: {str(e)}"}

# ------------------------------
# NextSLS API è°ƒç”¨å‡½æ•°ï¼ˆé€šç”¨ç‰ˆæœ¬ï¼‰
# ------------------------------
def call_nextsls(agent, tracking_number):
    """è°ƒç”¨NextSLS APIè·å–è½¨è¿¹ä¿¡æ¯ï¼ˆé€šç”¨ç‰ˆæœ¬ï¼‰"""
    try:
        # ä½¿ç”¨ä»£ç†é…ç½®çš„API URL
        url = agent.api_url
        if not url:
            return {"error": "æœªé…ç½®NextSLS APIåœ°å€"}
        
        # æ ¹æ®NextSLSæ–‡æ¡£æ„å»ºè¯·æ±‚å¤´
        headers = {
            "Content-Type": "application/json"
        }
        
        # æ·»åŠ è®¤è¯ä¿¡æ¯ï¼ˆæ ¹æ®ä¸åŒä»£ç†çš„é…ç½®æ–¹å¼ï¼‰
        if agent.app_token:
            headers["Authorization"] = f"Bearer {agent.app_token}"
        
        # æ„å»ºè¯·æ±‚ä½“ - æ”¯æŒå¤šç§æŸ¥è¯¢æ–¹å¼
        payload = {}
        
        # æ–¹å¼1: ä½¿ç”¨å®¢æˆ·å‚è€ƒå·ï¼ˆclient_referenceï¼‰æŸ¥è¯¢
        payload["shipment"] = {
            "client_reference": tracking_number,
            "language": "zh"
        }
        
        # æ–¹å¼2: å¦‚æœæœ‰access_tokenå‚æ•°ï¼Œæ·»åŠ åˆ°è¯·æ±‚ä¸­ï¼ˆæŸäº›NextSLSç‰ˆæœ¬éœ€è¦ï¼‰
        if agent.app_token and "access_token" not in payload:
            payload["access_token"] = agent.app_token
            
        # æ–¹å¼3: æŸäº›NextSLSç‰ˆæœ¬å¯èƒ½éœ€è¦ä¸åŒçš„å‚æ•°ç»“æ„
        # è¿™é‡Œå¯ä»¥æ ¹æ®ä»£ç†çš„å…·ä½“é…ç½®è¿›è¡Œè°ƒæ•´
        
        response = requests.post(url, headers=headers, json=payload, timeout=15)
        
        if response.status_code != 200:
            return {"error": f"NextSLS APIè¿”å›é”™è¯¯ {response.status_code}: {response.text}"}
            
        data = response.json()
        
        # æ£€æŸ¥APIå“åº”çŠ¶æ€ - æ”¯æŒå¤šç§å“åº”æ ¼å¼
        if data.get("status") not in [1, "1", 200, "200", True]:
            error_msg = data.get("info") or data.get("message") or data.get("error") or "æœªçŸ¥é”™è¯¯"
            return {"error": f"NextSLS APIé”™è¯¯: {error_msg}"}
            
        # è§£æè½¨è¿¹æ•°æ® - æ”¯æŒå¤šç§å“åº”æ ¼å¼
        tracks = []
        
        # æ ¼å¼1: data.shipment.traces (æ ‡å‡†NextSLSæ ¼å¼)
        if data.get("data") and data["data"].get("shipment"):
            shipment_data = data["data"]["shipment"]
            for trace in shipment_data.get("traces", []):
                # è½¬æ¢æ—¶é—´æˆ³ä¸ºæ—¥æœŸæ—¶é—´å­—ç¬¦ä¸²
                timestamp = trace.get("time")
                if timestamp and isinstance(timestamp, int):
                    time_str = datetime.fromtimestamp(timestamp).strftime("%Y-%m-%d %H:%M:%S")
                else:
                    time_str = trace.get("time", "")
                    
                tracks.append({
                    "track_occur_date": time_str,
                    "track_location": trace.get("location", ""),
                    "track_description": trace.get("info", trace.get("description", ""))
                })
        
        # æ ¼å¼2: ç›´æ¥åŒ…å«tracksæ•°ç»„
        elif data.get("tracks"):
            for trace in data.get("tracks", []):
                tracks.append({
                    "track_occur_date": trace.get("occur_date", trace.get("time", "")),
                    "track_location": trace.get("location", ""),
                    "track_description": trace.get("info", trace.get("description", ""))
                })
        
        # æ ¼å¼3: å…¶ä»–å¯èƒ½çš„æ ¼å¼
        elif data.get("data") and isinstance(data["data"], list):
            for item in data["data"]:
                tracks.append({
                    "track_occur_date": item.get("occur_date", item.get("time", "")),
                    "track_location": item.get("location", ""),
                    "track_description": item.get("info", item.get("description", ""))
                })
        
        # å¦‚æœæ²¡æœ‰æ‰¾åˆ°è½¨è¿¹æ•°æ®ï¼Œè¿”å›é”™è¯¯
        if not tracks:
            return {"error": "æœªæ‰¾åˆ°è½¨è¿¹ä¿¡æ¯"}
        
        return {
            "success": "1",
            "cnmessage": "NextSLSè½¨è¿¹æŸ¥è¯¢æˆåŠŸ",
            "data": [{"details": tracks}]
        }
        
    except Exception as e:
        return {"error": f"NextSLS APIè¯·æ±‚å¤±è´¥: {str(e)}"}

# ------------------------------
# å¤šè´§ä»£ APIï¼šgettrackï¼ˆTXFBA å¼ºåˆ¶ POST ç‰ˆæœ¬ï¼‰
# ------------------------------
def call_gettrack(carrier_id=None, tracking_number=None, agent_id=None, timeout=15):
    if not tracking_number:
        return {"error": "å¿…é¡»æä¾› tracking_number"}

    cache_key = (f"agent:{agent_id}" if agent_id else f"carrier:{carrier_id}", tracking_number)
    now = time.time()
    if cache_key in CACHE:
        ts, data = CACHE[cache_key]
        if now - ts < CACHE_TTL:
            return data

    # ------------------------------
    # æœ¬åœ°æ‰‹å·¥è½¨è¿¹
    # ------------------------------
    def local_manual():
        s = Shipment.query.filter_by(tracking_number=tracking_number).first()
        if not s or not s.manual_tracks:
            return None
        details = []
        for t in sorted(s.manual_tracks, key=lambda x: (x.happen_time or x.created_at or datetime.utcnow()), reverse=True):
            details.append({
                "track_occur_date": (t.happen_time or t.created_at).strftime("%Y-%m-%d %H:%M:%S"),
                "track_location": t.location or "",
                "track_description": t.description
            })
        return {
            "success": "1",
            "cnmessage": "æ‰‹å·¥è½¨è¿¹",
            "data": [{"details": details}]
        }

    # ------------------------------
    # è·å– shipment
    # ------------------------------
    shipment = Shipment.query.filter_by(tracking_number=tracking_number).first()
    if not shipment:
        return {"error": f"æœªæ‰¾åˆ°è¿å•å·: {tracking_number}"}

    shipment_id_to_use = shipment.agent_tracking_number
    if not shipment_id_to_use:
        return {"error": f"æœªæ‰¾åˆ°å¯¹åº”çš„ shipment_idï¼ˆagent_tracking_numberï¼‰: {tracking_number}"}

    # ------------------------------
    # ä½¿ç”¨ DB agent
    # ------------------------------
    if agent_id or shipment.agent_id:
        agent = CarrierAgent.query.get(int(agent_id or shipment.agent_id))
        if not agent or not agent.is_active:
            data = {"error": "æœªæ‰¾åˆ°æŒ‡å®šä»£ç†æˆ–å·²åœç”¨"}
            CACHE[cache_key] = (now, data)
            return data

        if not agent.supports_api:
            data = local_manual() or {"error": "è¯¥ä»£ç†ä¸æ”¯æŒæŠ“å–ï¼Œä¸”æ— æ‰‹å·¥è½¨è¿¹"}
            CACHE[cache_key] = (now, data)
            return data

        # ----------------
        # NextSLS é€»è¾‘
        # ----------------
        is_nextsls = "nextsls" in (agent.api_url or "").lower() or "sls" in (agent.api_url or "").lower()
        if is_nextsls:
            payload = {
                "shipment": {
                    "shipment_id": shipment_id_to_use,
                    "language": "zh"
                }
            }
            headers = {
                "Content-Type": "application/json",
                "Authorization": f"Bearer {agent.app_token or API_TOKEN}"
            }
            try:
                r = requests.post(agent.api_url or API_URL, json=payload, headers=headers, timeout=timeout)
                r.encoding = "utf-8"
                data = r.json()
                if data.get("status") != 1:
                    data["hint"] = f"ç–‘ä¼¼ shipment_id é”™è¯¯: {shipment_id_to_use}"
                CACHE[cache_key] = (now, data)
                return data
            except Exception as e:
                data = {"error": f"è¯·æ±‚ä»£ç†æ¥å£å‡ºé”™: {e}"}
                CACHE[cache_key] = (now, data)
                return data

        # ----------------
        # TXFBA å¼ºåˆ¶ POST
        # ----------------
        if "txfba.com" in (agent.api_url or "").lower():
            payload = {
                "appToken": agent.app_token or "",
                "appKey": agent.app_key or "",
                "serviceMethod": "gettrack",
                "paramsJson": json.dumps({"tracking_number": tracking_number}, ensure_ascii=False)
            }
            headers = {"Content-Type": "application/x-www-form-urlencoded"}

            # ğŸ”¹ è°ƒè¯•æ‰“å°
            print("TXFBA è¯·æ±‚ URL:", agent.api_url)
            print("è¯·æ±‚ payload:", payload)
            print("è¯·æ±‚ headers:", headers)

            try:
                r = requests.post(agent.api_url, data=payload, headers=headers, timeout=timeout)
                r.encoding = "utf-8"
                try:
                    data = r.json()
                except Exception:
                    data = {"raw_text": r.text}
                if isinstance(data, dict) and data.get("success") == "0" and "appToken" in (data.get("cnmessage") or ""):
                    data["hint"] = "ç–‘ä¼¼ appKey/appToken æˆ–å®¢æˆ·å·é…ç½®é”™è¯¯ï¼Œæˆ–è¯¥ä»£ç†æœªå¼€é€š API æƒé™"
                CACHE[cache_key] = (now, data)
                return data
            except Exception as e:
                data = {"error": f"è¯·æ±‚ä»£ç†æ¥å£å‡ºé”™: {e}"}
                CACHE[cache_key] = (now, data)
                return data

        # ----------------
        # RTB56 / å…¶ä»–ä»£ç†é€»è¾‘
        # ----------------
        if agent.app_key or agent.app_token:
            try:
                payload = {
                    "appToken": agent.app_token or "",
                    "appKey": agent.app_key or "",
                    "serviceMethod": "gettrack",
                    "paramsJson": json.dumps({"tracking_number": tracking_number}, ensure_ascii=False)
                }
                headers = {"Content-Type": "application/x-www-form-urlencoded"}
                r = requests.post(agent.api_url, data=payload, headers=headers, timeout=timeout)
                r.encoding = "utf-8"
                try:
                    data = r.json()
                except Exception:
                    data = {"raw_text": r.text}
                CACHE[cache_key] = (now, data)
                return data
            except Exception as e:
                data = {"error": f"è¯·æ±‚ä»£ç†æ¥å£å‡ºé”™: {e}"}
                CACHE[cache_key] = (now, data)
                return data

        # ----------------
        # ç”¨æˆ·å/å¯†ç æ–¹å¼
        # ----------------
        try:
            payload = {
                "username": agent.username or "",
                "password": agent.password or "",
                "tracking_number": tracking_number
            }
            r = requests.post(agent.api_url, data=payload, timeout=timeout)
            r.encoding = "utf-8"
            try:
                data = r.json()
            except Exception:
                data = {"raw_text": r.text}
            CACHE[cache_key] = (now, data)
            return data
        except Exception as e:
            data = {"error": f"è¯·æ±‚ä»£ç†æ¥å£å‡ºé”™: {e}"}
            CACHE[cache_key] = (now, data)
            return data

    # ----------------
    # ä½¿ç”¨ç¯å¢ƒ carrier
    # ----------------
    carrier = CARRIERS.get(carrier_id) if carrier_id else CARRIERS.get(CARRIERS_LIST[0]) if CARRIERS_LIST else None
    if not carrier:
        data = {"error": "æœªé…ç½®å¯ç”¨çš„è´§ä»£ï¼ˆenvï¼‰"}
        CACHE[cache_key] = (now, data)
        return data

    payload = {
        "appToken": carrier.get("token"),
        "appKey": carrier.get("key"),
        "serviceMethod": "gettrack",
        "paramsJson": json.dumps({"tracking_number": tracking_number}, ensure_ascii=False)
    }
    headers = {"Content-Type": "application/x-www-form-urlencoded"}
    try:
        r = requests.post(carrier.get("api_url", DEFAULT_API_URL), data=payload, headers=headers, timeout=timeout)
        r.encoding = "utf-8"
        try:
            data = r.json()
        except Exception:
            data = {"raw_text": r.text}
        if isinstance(data, dict) and data.get("success") == "0":
            local = local_manual()
            if local:
                data = local
        CACHE[cache_key] = (now, data)
        return data
    except Exception as e:
        data = local_manual() or {"error": f"è¯·æ±‚å‡ºé”™: {e}"}
        CACHE[cache_key] = (now, data)
        return data

# ------------------------------
# è½¨è¿¹æ ¼å¼åŒ–
# ------------------------------
def format_tracks_from_data(data):
    if not data:
        return "æ²¡æœ‰è¿”å›æ•°æ®"
    if isinstance(data, dict) and data.get("error"):
        return f"é”™è¯¯: {data.get('error')}"
    if isinstance(data, dict) and "raw_text" in data:
        return data["raw_text"][:10000]
    if isinstance(data, dict) and "data" in data and isinstance(data["data"], list) and data["data"]:
        first = data["data"][0]
        details = first.get("details") or []
        if not details:
            return data.get("cnmessage", "æš‚æ— è½¨è¿¹ä¿¡æ¯")
        parts = []
        for d in details:
            t = (d.get("track_occur_date") or "").strip()
            loc = (d.get("track_location") or "").strip()
            desc = (d.get("track_description") or d.get("track_description_en") or "").strip()
            line = " â€” ".join([x for x in [loc, desc] if x])
            parts.append(f"{line}\n{t}".strip())
        return "\n\n".join(parts) if parts else "æš‚æ— è½¨è¿¹ä¿¡æ¯"
    try:
        return json.dumps(data, ensure_ascii=False, indent=2)
    except Exception:
        return str(data)

# ------------------------------
# Excel å¯¼å‡º
# ------------------------------
def generate_invoice_xlsx(company_name, customer_name, bank_info, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "è´¦å•"

    ws.merge_cells("A1:I1")
    ws["A1"] = company_name

    ws["A3"] = "å®¢æˆ·åç§°:"
    ws["B3"] = customer_name

    headers = ["åºå·", "æ—¥æœŸ", "è®¢å•å·", "æœåŠ¡å•†å•å·", "ä»¶æ•°", "è®¡è´¹é‡/KG", "ç›®çš„åœ°", "è¿è¾“æ¸ é“", "åˆè®¡è´¹ç”¨", "è´¦å•æ‘˜è¦"]
    ws.append(headers)

    start_row = ws.max_row + 1
    for idx, r in enumerate(rows, start=1):
        r2 = r + [""] * (9 - len(r))
        ws.append([idx] + r2[:9])

    total = 0.0
    data_start = start_row
    for i in range(len(rows)):
        try:
            cell = ws.cell(row=data_start + i, column=9)
            if cell.value is not None:
                total += float(str(cell.value))
        except Exception:
            pass

    summary_row = data_start + len(rows) + 1
    ws[f"H{summary_row}"] = "åˆè®¡è´¹ç”¨"
    ws[f"I{summary_row}"] = total

    info_row = summary_row + 2
    ws[f"A{info_row}"] = "æ”¶æ¬¾é“¶è¡Œä¿¡æ¯"
    ws[f"A{info_row+1}"] = bank_info

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value else ""
                if len(val) > max_length:
                    max_length = len(val)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

# ------------------------------
# ç™»å½•/ç™»å‡º
# ------------------------------
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        if not username or not password:
            flash("è¯·è¾“å…¥ç”¨æˆ·åä¸å¯†ç ")
            return redirect(url_for("login"))
        user = User.query.filter_by(username=username).first()
        if not user or not user.check_password(password):
            flash("ç”¨æˆ·åæˆ–å¯†ç é”™è¯¯")
            return redirect(url_for("login"))
        login_user(user)
        flash("ç™»å½•æˆåŠŸ")
        return redirect(url_for("index"))
    try:
        return render_template("login.html")
    except TemplateNotFound:
        return render_template_string("""
        <!doctype html><meta charset="utf-8"><title>ç™»å½•</title>
        <h2>ç™»å½•</h2>
        <form method="post">
            ç”¨æˆ·å: <input name="username"><br><br>
            å¯†ç : <input name="password" type="password"><br><br>
            <button type="submit">ç™»å½•</button>
        </form>
        """)

@app.route("/logout")
@login_required
def logout():
    logout_user()
    flash("å·²ç™»å‡º")
    return redirect(url_for("login"))

# ------------------------------
# é¦–é¡µ
# ------------------------------
@app.route("/")
@login_required
def index():
    return redirect(url_for("shipments"))

# ------------------------------
# ä»£ç†ç®¡ç†ï¼šæ–°å¢ / åˆ—è¡¨ï¼ˆå…¼å®¹ä½ çš„æ¨¡æ¿ï¼‰
# æ–°å¢ï¼šç¼–è¾‘ã€è½¯åˆ é™¤
# ------------------------------
@app.route("/agents", methods=["GET", "POST"])
@login_required
def agents():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        api_url = request.form.get("api_url", "").strip()
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        app_key = request.form.get("app_key", "").strip()
        app_token = request.form.get("app_token", "").strip()
        customer_code = request.form.get("customer_code", "").strip()
        if not name:
            flash("ä»£ç†åç§°ä¸èƒ½ä¸ºç©º")
            return redirect(url_for("agents"))
        a = CarrierAgent(
            name=name, api_url=api_url, username=username, password=password,
            app_key=app_key, app_token=app_token, customer_code=customer_code,
            is_active=True
        )
        db.session.add(a)
        db.session.commit()
        flash("ä»£ç†å·²ä¿å­˜")
        return redirect(url_for("agents"))
    data = CarrierAgent.query.filter_by(is_active=True).order_by(CarrierAgent.name).all()
    return render_template_safe("agent.html", agents=data)

@app.route("/agents/<int:agent_id>/edit", methods=["GET", "POST"])
@login_required
def edit_agent(agent_id):
    a = CarrierAgent.query.get_or_404(agent_id)
    if not a.is_active:
        flash("è¯¥ä»£ç†å·²åœç”¨")
        return redirect(url_for("agents"))
    if request.method == "POST":
        a.name = request.form.get("name", a.name).strip() or a.name
        a.api_url = request.form.get("api_url", a.api_url).strip()
        a.username = request.form.get("username", a.username).strip()
        a.password = request.form.get("password", a.password).strip()
        a.app_key = request.form.get("app_key", a.app_key).strip()
        a.app_token = request.form.get("app_token", a.app_token).strip()
        a.customer_code = request.form.get("customer_code", a.customer_code).strip()
        a.supports_api = (request.form.get("supports_api", "1") == "1")
        db.session.commit()
        flash("ä»£ç†å·²æ›´æ–°")
        return redirect(url_for("agents"))
    # æä¾›ä¸€ä¸ªç®€å•å ä½ç¼–è¾‘è¡¨å•ï¼ˆè‹¥æ²¡æœ‰å‰ç«¯æ¨¡æ¿ï¼‰
    return render_template_string("""
    <h3>ç¼–è¾‘ä»£ç†</h3>
    <form method="post">
      åç§° <input name="name" value="{{a.name}}"><br>
      API  <input name="api_url" value="{{a.api_url or ''}}"><br>
      è´¦å· <input name="username" value="{{a.username or ''}}"><br>
      å¯†ç  <input name="password" value="{{a.password or ''}}"><br>
      appKey <input name="app_key" value="{{a.app_key or ''}}"><br>
      appToken <input name="app_token" value="{{a.app_token or ''}}"><br>
      å®¢æˆ·å· <input name="customer_code" value="{{a.customer_code or ''}}"><br>
      æ”¯æŒAPI <select name="supports_api"><option value="1" {% if a.supports_api %}selected{% endif %}>æ˜¯</option>
      <option value="0" {% if not a.supports_api %}selected{% endif %}>å¦</option></select><br><br>
      <button>ä¿å­˜</button> <a href="{{url_for('agents')}}">è¿”å›</a>
    </form>
    """, a=a)

@app.route("/agents/<int:agent_id>/delete", methods=["POST"])
@login_required
def delete_agent(agent_id):
    a = CarrierAgent.query.get_or_404(agent_id)
    # è½¯åˆ é™¤ï¼šè§£é™¤è¿å•å…³è”ï¼Œé¿å…å¤–é”®é”™è¯¯
    Shipment.query.filter_by(agent_id=a.id).update({"agent_id": None})
    a.is_active = False
    db.session.commit()
    flash("å·²åœç”¨è¯¥ä»£ç†ï¼ˆè½¯åˆ é™¤ï¼‰ï¼Œå¹¶è§£ç»‘ç›¸å…³è¿å•ã€‚")
    return redirect(url_for("agents"))

# ------------------------------
# è¿å•ç®¡ç†ï¼šåˆ—è¡¨ / æ–°å¢ / å¯¼å…¥
# ------------------------------
@app.route("/shipments")
@login_required
def shipments():
    data = Shipment.query.order_by(Shipment.created_at.desc()).all()
    # å…¼å®¹ä½ çš„æ¨¡æ¿ï¼ˆshipments.html é‡Œç›´æ¥æ¸²æŸ“å·²æœ‰ dataï¼‰
    return render_template_safe("shipments.html",
                                shipments=data,
                                customers=Customer.query.order_by(Customer.name).all(),
                                agents=CarrierAgent.query.filter_by(is_active=True).order_by(CarrierAgent.name).all(),
                                destinations=["ç¾å›½", "é¦™æ¸¯", "ä¸­å›½", "è‹±å›½", "å¾·å›½", "å…¶ä»–"])

def _calc_fee(weight, unit_price, surcharge_extra, operation_fee, high_value_fee, manual_fee):
    try:
        w = float(weight or 0)
        p = float(unit_price or 0)
        s1 = float(surcharge_extra or 0)
        s2 = float(operation_fee or 0)
        s3 = float(high_value_fee or 0)
        if manual_fee is not None and str(manual_fee).strip() != "":
            return float(manual_fee)
        return round(w * p + s1 + s2 + s3, 2)
    except Exception:
        return float(manual_fee or 0)

@app.route("/shipments/add", methods=["GET", "POST"])
@login_required
def add_shipment():
    customers = Customer.query.order_by(Customer.name).all()
    agents = CarrierAgent.query.filter_by(is_active=True).order_by(CarrierAgent.name).all()
    carriers = CARRIERS
    destinations = ["ç¾å›½", "é¦™æ¸¯", "ä¸­å›½", "è‹±å›½", "å¾·å›½", "å…¶ä»–"]
    if request.method == "POST":
        tn = request.form.get("tracking_number", "").strip()
        if not tn:
            flash("è¯·å¡«å†™è¿å•å·")
            return redirect(url_for("add_shipment"))

        # è·å– shipment_id
        shipment_id = request.form.get("shipment_id", "").strip()
        third_party_tracking_number = request.form.get("third_party_tracking_number", "").strip()

        # å¯é€‰å­—æ®µ
        customer_id = request.form.get("customer_id") or None
        agent_id = request.form.get("agent_id") or None
        carrier_id = request.form.get("carrier_id") or None
        destination = request.form.get("destination", "").strip()
        channel = request.form.get("channel", "").strip()
        product_type = request.form.get("product_type", "").strip()
        pieces = int(request.form.get("pieces", "1") or 1)

        weight = request.form.get("weight")
        unit_price = request.form.get("unit_price")
        surcharge_extra = request.form.get("surcharge_extra")
        operation_fee = request.form.get("operation_fee")
        high_value_fee = request.form.get("high_value_fee")

        manual_fee = request.form.get("fee")  # å¦‚æœä½ ä»åœ¨é¡µé¢ä¸Šæ‰‹å¡«åˆè®¡ï¼Œè¿™é‡Œä¼˜å…ˆç”Ÿæ•ˆ
        fee = _calc_fee(weight, unit_price, surcharge_extra, operation_fee, high_value_fee, manual_fee)

        note = request.form.get("note", "")

        s = Shipment(
            tracking_number=tn,
            shipment_id=shipment_id,  # ä½¿ç”¨ shipment_id
            third_party_tracking_number=third_party_tracking_number,
            carrier_id=carrier_id,
            agent_id=int(agent_id) if agent_id else None,
            customer_id=int(customer_id) if customer_id else None,
            origin=request.form.get("origin", ""),
            destination=destination,
            channel=channel,
            product_type=product_type,
            pieces=pieces,
            weight=float(weight or 0),
            unit_price=float(unit_price or 0),
            surcharge_extra=float(surcharge_extra or 0),
            operation_fee=float(operation_fee or 0),
            high_value_fee=float(high_value_fee or 0),
            fee=fee,
            note=note,
            status="å·²å½•å…¥"
        )
        db.session.add(s)
        db.session.commit()
        flash("è¿å•å·²ä¿å­˜")
        return redirect(url_for("shipments"))
    return render_template_safe("add_shipment.html",
                                customers=customers, agents=agents, carriers=carriers, destinations=destinations)

@app.route("/shipments/import", methods=["POST"])
@login_required
def import_shipments():
    if pd is None:
        flash("æœªå®‰è£… pandasï¼Œæ— æ³•ä½¿ç”¨ Excel å¯¼å…¥ã€‚è¯·å®‰è£… pandas")
        return redirect(url_for("shipments"))
    f = request.files.get("file")
    if not f:
        flash("æœªä¸Šä¼ æ–‡ä»¶")
        return redirect(url_for("shipments"))
    try:
        df = pd.read_excel(f)
    except Exception as e:
        flash(f"è¯»å– Excel å‡ºé”™: {e}")
        return redirect(url_for("shipments"))
    count = 0
    for _, row in df.iterrows():
        tn = str(row.get("tracking_number", "")).strip()
        if not tn or Shipment.query.filter_by(tracking_number=tn).first():
            continue
        s = Shipment(
            tracking_number=tn,
            origin=row.get("origin", ""),
            destination=row.get("destination", ""),
            weight=float(row.get("weight") or 0),
            unit_price=float(row.get("unit_price") or 0),
            surcharge_extra=float(row.get("surcharge_extra") or 0),
            operation_fee=float(row.get("operation_fee") or 0),
            high_value_fee=float(row.get("high_value_fee") or 0),
            fee=_calc_fee(row.get("weight"), row.get("unit_price"),
                          row.get("surcharge_extra"), row.get("operation_fee"),
                          row.get("high_value_fee"), row.get("fee")),
            status="å·²å¯¼å…¥"
        )
        db.session.add(s)
        count += 1
    db.session.commit()
    flash(f"å…±å¯¼å…¥ {count} æ¡è¿å•")
    return redirect(url_for("shipments"))

# é¢å¤–ï¼šè¿å•ç¼–è¾‘/åˆ é™¤ï¼ˆé¿å…è¯¯å½•æ— æ³•æ”¹ï¼‰
@app.route("/shipments/<int:sid>/edit", methods=["GET", "POST"])
@login_required
def edit_shipment(sid):
    s = Shipment.query.get_or_404(sid)
    if request.method == "POST":
        # æ›´æ–°æ–°å­—æ®µ
        s.agent_tracking_number = request.form.get("agent_tracking_number", s.agent_tracking_number)
        s.third_party_tracking_number = request.form.get("third_party_tracking_number", s.third_party_tracking_number)
        
        for f in ["destination", "channel", "product_type", "note", "status", "origin"]:
            setattr(s, f, request.form.get(f, getattr(s, f)))
        s.customer_id = int(request.form.get("customer_id")) if request.form.get("customer_id") else s.customer_id
        s.agent_id = int(request.form.get("agent_id")) if request.form.get("agent_id") else s.agent_id
        s.pieces = int(request.form.get("pieces", s.pieces) or s.pieces)
        s.weight = float(request.form.get("weight", s.weight) or s.weight)
        s.unit_price = float(request.form.get("unit_price", s.unit_price) or s.unit_price)
        s.surcharge_extra = float(request.form.get("surcharge_extra", s.surcharge_extra) or s.surcharge_extra)
        s.operation_fee = float(request.form.get("operation_fee", s.operation_fee) or s.operation_fee)
        s.high_value_fee = float(request.form.get("high_value_fee", s.high_value_fee) or s.high_value_fee)
        manual_fee = request.form.get("fee")
        s.fee = _calc_fee(s.weight, s.unit_price, s.surcharge_extra, s.operation_fee, s.high_value_fee, manual_fee)
        db.session.commit()
        flash("è¿å•å·²æ›´æ–°")
        return redirect(url_for("shipments"))
    return render_template_string("""
    <h3>ç¼–è¾‘è¿å• {{s.tracking_number}}</h3>
    <form method="post">
      ä»£ç†ç³»ç»Ÿå•å· <input name="agent_tracking_number" value="{{s.agent_tracking_number or ''}}"><br>
      17Trackå•å· <input name="third_party_tracking_number" value="{{s.third_party_tracking_number or ''}}"><br>
      ç›®çš„åœ° <input name="destination" value="{{s.destination or ''}}"><br>
      æ¸ é“ <input name="channel" value="{{s.channel or ''}}"><br>
      äº§å“ç±»å‹ <input name="product_type" value="{{s.product_type or ''}}"><br>
      ä»¶æ•° <input name="pieces" value="{{s.pieces or 1}}"><br>
      é‡é‡ <input name="weight" value="{{s.weight or 0}}"><br>
      å•ä»· <input name="unit_price" value="{{s.unit_price or 0}}"><br>
      é™„åŠ è´¹ <input name="surcharge_extra" value="{{s.surcharge_extra or 0}}"><br>
      æ“ä½œè´¹ <input name="operation_fee" value="{{s.operation_fee or 0}}"><br>
      è¶…å€¼è´¹ <input name="high_value_fee" value="{{s.high_value_fee or 0}}"><br>
      åˆè®¡(å¯æ‰‹å¡«è¦†ç›–) <input name="fee" value="{{s.fee or 0}}"><br>
      å¤‡æ³¨ <input name="note" value="{{s.note or ''}}"><br><br>
      <button>ä¿å­˜</button> <a href="{{url_for('shipments')}}">è¿”å›</a>
    </form>
    """, s=s)

@app.route("/shipments/<int:sid>/delete", methods=["POST"])
@login_required
def delete_shipment(sid):
    s = Shipment.query.get_or_404(sid)
    # å…è®¸åˆ é™¤ï¼ˆæ— å¤–é”®çº¦æŸå½±å“å…¶ä»–è¡¨ï¼‰
    ManualTrack.query.filter_by(shipment_id=s.id).delete()
    db.session.delete(s)
    db.session.commit()
    flash("è¿å•å·²åˆ é™¤")
    return redirect(url_for("shipments"))

# ------------------------------
# å®¢æˆ·ç®¡ç†ï¼ˆå…¼å®¹åŸæœ‰æ¨¡æ¿ï¼‰
# ------------------------------
@app.route("/customers", methods=["GET", "POST"])
@login_required
def customers():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        bankinfo = request.form.get("bankinfo", "").strip()
        if not name:
            flash("å®¢æˆ·åç§°ä¸èƒ½ä¸ºç©º")
            return redirect(url_for("customers"))
        c = Customer(name=name, bank_info=bankinfo)
        db.session.add(c)
        db.session.commit()
        flash("å®¢æˆ·å·²ä¿å­˜")
        return redirect(url_for("customers"))
    data = Customer.query.order_by(Customer.name).all()
    return render_template_safe("customer.html", customers=data)

# ------------------------------
# è´¢åŠ¡æ¨¡å—ï¼ˆæ—¶é—´ç­›é€‰ + é€‰ç”¨æ”¶æ¬¾è´¦æˆ· + å¯¼å‡ºï¼‰
# ------------------------------
@app.route("/finance", methods=["GET", "POST"])
@login_required
def finance():
    total_shipments = Shipment.query.count()
    total_fees = db.session.query(db.func.sum(Shipment.fee)).scalar() or 0.0

    customers = Customer.query.order_by(Customer.name).all()
    accounts = BankAccount.query.order_by(BankAccount.is_default.desc(), BankAccount.id.desc()).all()

    selected_customer_id = None
    shipments = []
    date_from = None
    date_to = None
    bank_account_id = None

    if request.method == "POST":
        selected_customer_id = request.form.get("customer_id") or None
        date_from_raw = request.form.get("date_from") or ""
        date_to_raw = request.form.get("date_to") or ""
        bank_account_id = request.form.get("bank_account_id") or None

        q = Shipment.query
        if selected_customer_id:
            q = q.filter(Shipment.customer_id == int(selected_customer_id))
        if date_from_raw:
            try:
                date_from = datetime.strptime(date_from_raw + " 00:00:00", "%Y-%m-%d %H:%M:%S")
                q = q.filter(Shipment.created_at >= date_from)
            except Exception:
                pass
        if date_to_raw:
            try:
                date_to = datetime.strptime(date_to_raw + " 23:59:59", "%Y-%m-%d %H:%M:%S")
                q = q.filter(Shipment.created_at <= date_to)
            except Exception:
                pass

        shipments = q.order_by(Shipment.created_at.desc()).all()

        if request.form.get("export") == "1":
            rows = []
            company_name = os.getenv("COMPANY_NAME", "å…¬å¸åç§°")
            customer = Customer.query.get(int(selected_customer_id)) if selected_customer_id else None

            # é€‰ä¸­çš„æ”¶æ¬¾è´¦æˆ·æ–‡æœ¬
            bank_info_text = ""
            if bank_account_id:
                acc = BankAccount.query.get(int(bank_account_id))
                if acc:
                    bank_info_text = f"{'å…¬è´¦' if acc.account_type=='public' else 'ç§äºº'}\nå¼€æˆ·è¡Œï¼š{acc.bank_name}\næˆ·åï¼š{acc.account_name}\nè´¦å·ï¼š{acc.account_no}"
            else:
                # ä½¿ç”¨å®¢æˆ·é»˜è®¤ bank_info æˆ–é»˜è®¤è´¦æˆ·
                if customer and customer.bank_info:
                    bank_info_text = customer.bank_info
                else:
                    acc = BankAccount.query.filter_by(is_default=True).first()
                    if acc:
                        bank_info_text = f"{'å…¬è´¦' if acc.account_type=='public' else 'ç§äºº'}\nå¼€æˆ·è¡Œï¼š{acc.bank_name}\næˆ·åï¼š{acc.account_name}\nè´¦å·ï¼š{acc.account_no}"

            for s in shipments:
                rows.append([
                    s.created_at.strftime("%Y-%m-%d") if s.created_at else "",
                    "", s.tracking_number, "",
                    (s.pieces or 1), (s.weight or 0),
                    s.destination or "", s.channel or "",
                    (s.fee or 0), s.note or ""
                ])
            bio = generate_invoice_xlsx(company_name, customer.name if customer else "å®¢æˆ·", bank_info_text or "", rows)
            filename = f"invoice_{customer.name if customer else 'customer'}_{int(time.time())}.xlsx"
            return send_file(bio, as_attachment=True, download_name=filename,
                             mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    return render_template_safe("finance.html",
                                total_shipments=total_shipments,
                                total_fees=total_fees,
                                customers=customers,
                                shipments=shipments,
                                selected_customer_id=selected_customer_id,
                                accounts=accounts,
                                bank_account_id=bank_account_id,
                                date_from=date_from.strftime("%Y-%m-%d") if date_from else "",
                                date_to=date_to.strftime("%Y-%m-%d") if date_to else "")

# ç®€æ˜“é“¶è¡Œè´¦æˆ·ç®¡ç†ï¼ˆå ä½è·¯ç”±ï¼Œå¯åœ¨ Admin é‡Œç›´æ¥ç»´æŠ¤ï¼‰
@app.route("/bank_accounts", methods=["GET", "POST"])
@login_required
def bank_accounts():
    if request.method == "POST":
        acc = BankAccount(
            account_type=request.form.get("account_type", "private"),
            bank_name=request.form.get("bank_name", "").strip(),
            account_name=request.form.get("account_name", "").strip(),
            account_no=request.form.get("account_no", "").strip(),
            is_default=(request.form.get("is_default") == "1"),
            remark=request.form.get("remark", "").strip(),
        )
        if acc.is_default:
            BankAccount.query.update({"is_default": False})
        db.session.add(acc)
        db.session.commit()
        flash("å·²æ–°å¢æ”¶æ¬¾è´¦æˆ·")
        return redirect(url_for("bank_accounts"))
    data = BankAccount.query.order_by(BankAccount.is_default.desc(), BankAccount.id.desc()).all()
    return render_template_string("""
    <h3>æ”¶æ¬¾è´¦æˆ·</h3>
    <form method="post">
      ç±»å‹ <select name="account_type"><option value="private">ç§äºº</option><option value="public">å…¬è´¦</option></select><br>
      å¼€æˆ·è¡Œ <input name="bank_name"><br>
      æˆ·å <input name="account_name"><br>
      è´¦å· <input name="account_no"><br>
      é»˜è®¤ <input type="checkbox" name="is_default" value="1"><br>
      å¤‡æ³¨ <input name="remark"><br><br>
      <button>æ–°å¢</button>
    </form>
    <hr>
    <ul>
      {% for a in data %}
      <li>[{{'é»˜è®¤' if a.is_default else ' '}}] {{a.bank_name}} / {{a.account_name}} / {{a.account_no}} ({{'å…¬è´¦' if a.account_type=='public' else 'ç§äºº'}})</li>
      {% endfor %}
    </ul>
    """, data=data)

# ------------------------------
# æ‰‹å·¥è½¨è¿¹ï¼šä¸ºä¸æ”¯æŒæŠ“å–çš„ä»£ç†/å•ç¥¨æ·»åŠ 
# ------------------------------
from datetime import datetime
from flask import request, redirect, url_for, flash, render_template_string, jsonify
from flask_login import login_required, current_user

# æ³¨æ„ï¼šutils.py å¦‚æœå­˜åœ¨ call_gettrack ç­‰ï¼Œéœ€ç¡®ä¿å¯¼å…¥æ­£ç¡®
# å‡è®¾ utils.py æœ‰è¿™äº›å‡½æ•°
try:
    from utils import call_gettrack, format_tracks_from_data
except ImportError:
    pass  # å¦‚æœæ²¡æœ‰ï¼Œå¿½ç•¥æˆ–å®šä¹‰

# ------------------------------
# æ‰‹å·¥è½¨è¿¹
# ------------------------------
@app.route("/shipments/<int:sid>/tracks", methods=["GET", "POST"])
@login_required
def manual_tracks(sid):
    s = Shipment.query.get_or_404(sid)
    if request.method == "POST":
        desc = request.form.get("description", "").strip()
        if not desc:
            flash("è¯·è¾“å…¥è½¨è¿¹æè¿°")
            return redirect(url_for("manual_tracks", sid=sid))
        t = ManualTrack(
            shipment_id=s.id,
            happen_time=datetime.strptime(request.form.get("happen_time"), "%Y-%m-%d %H:%M:%S") if request.form.get("happen_time") else datetime.utcnow(),
            location=request.form.get("location", "").strip(),
            description=desc,
            author=current_user.username if current_user.is_authenticated else None
        )
        db.session.add(t)
        db.session.commit()
        flash("å·²ä¿å­˜æ‰‹å·¥è½¨è¿¹")
        return redirect(url_for("manual_tracks", sid=sid))
    return render_template_string("""
    <h3>æ‰‹å·¥è½¨è¿¹ - {{s.tracking_number}}</h3>
    <form method="post">
      æ—¶é—´(YYYY-MM-DD HH:MM:SS) <input name="happen_time"><br>
      åœ°ç‚¹ <input name="location"><br>
      æè¿° <textarea name="description" rows="4" cols="60"></textarea><br><br>
      <button>æ·»åŠ </button> <a href="{{url_for('shipments')}}">è¿”å›</a>
    </form>
    <hr>
    <pre style="white-space:pre-wrap">
    {% for t in s.manual_tracks|sort(attribute='happen_time', reverse=True) %}
    {{ (t.happen_time or t.created_at).strftime('%Y-%m-%d %H:%M:%S') }}  {{ t.location or '' }}
    {{ t.description }}
    ---------------------------
    {% endfor %}
    </pre>
    """, s=s)

# ------------------------------
# å†…éƒ¨è½¨è¿¹æŸ¥è¯¢ï¼ˆæ”¯æŒä»£ç†/å®¢æˆ·ï¼‰
# ------------------------------
@app.route("/track", methods=["GET", "POST"])
@login_required
def track():
    results = {}
    default_text = ""
    message = ""
    agents = CarrierAgent.query.filter_by(is_active=True).order_by(CarrierAgent.name).all()
    carriers_env = CARRIERS
    if request.method == "POST":
        forced_carrier = request.form.get("carrier_id") or None
        agent_id = request.form.get("agent_id") or None
        customer_id = request.form.get("customer_id") or None
        default_text = request.form.get("numbers", "").strip()

        numbers = []
        if default_text:
            numbers = [ln.strip() for ln in default_text.splitlines() if ln.strip()]

        if not numbers:
            q = Shipment.query
            if agent_id:
                q = q.filter(Shipment.agent_id == int(agent_id))
            if customer_id:
                q = q.filter(Shipment.customer_id == int(customer_id))
            numbers = [s.tracking_number for s in q.order_by(Shipment.created_at.desc()).limit(30).all()]
            if not numbers:
                message = "æœªæ‰¾åˆ°ç¬¦åˆæ¡ä»¶çš„è¿å•ã€‚"
        else:
            if len(numbers) > 30:
                message = f"è¾“å…¥ {len(numbers)} æ¡ï¼Œæœ¬æ¬¡åªå¤„ç†å‰ 30 æ¡ã€‚"
                numbers = numbers[:30]

        for n in numbers:
            shipment = Shipment.query.filter_by(tracking_number=n).first()
            if shipment and shipment.agent_tracking_number:
                data = call_gettrack(None, n, agent_id=agent_id or shipment.agent_id)
            else:
                data = {"error": f"æœªæ‰¾åˆ°å¯¹åº”çš„ shipment_idï¼ˆagent_tracking_numberï¼‰: {n}"}
            if isinstance(data, dict) and data.get("error"):
                results[n] = {"error": data.get("error"), "tracks": None, "raw": data}
            else:
                results[n] = {"error": None, "tracks": format_tracks_from_data(data), "raw": data}

    return render_template_safe("track.html",
                                carriers=carriers_env, agents=agents,
                                customers=Customer.query.order_by(Customer.name).all(),
                                results=results, message=message, default_text=default_text)

# ------------------------------
# å…¬å…±æŸ¥è¯¢é¡µé¢ï¼ˆWebè¡¨å•ï¼‰
# ------------------------------
@app.route("/public/track", methods=["GET", "POST"])
def public_track_page():
    results = {}
    default_text = ""
    message = ""
    if request.method == "POST":
        default_text = request.form.get("numbers", "").strip()
        if not default_text:
            message = "è¯·è¾“å…¥è¿å•å·"
        else:
            lines = [ln.strip() for ln in default_text.splitlines() if ln.strip()]
            if len(lines) > 30:
                message = f"è¾“å…¥ {len(lines)} æ¡ï¼Œæœ¬æ¬¡åªå¤„ç†å‰ 30 æ¡ã€‚"
                lines = lines[:30]
            for n in lines:
                s = Shipment.query.filter_by(tracking_number=n).first()
                data = None
                
                # ä¼˜å…ˆå°è¯•17TrackæŸ¥è¯¢ï¼ˆå¦‚æœæœ‰17Trackå•å·ï¼‰
                if s and s.third_party_tracking_number:
                    data = call_17track(s.third_party_tracking_number)
                
                # å¦‚æœæ²¡æœ‰17Trackç»“æœï¼Œå°è¯•å¸¸è§„æŸ¥è¯¢
                if not data or (isinstance(data, dict) and data.get("error")):
                    if s and s.agent_id:
                        data = call_gettrack(None, n, agent_id=s.agent_id)
                    elif s and s.carrier_id:
                        data = call_gettrack(s.carrier_id, n, agent_id=None)
                    else:
                        data = call_gettrack(None, n, agent_id=None)
                
                if isinstance(data, dict) and data.get("error"):
                    results[n] = {"error": data.get("error"), "tracks": None, "raw": data}
                else:
                    results[n] = {"error": None, "tracks": format_tracks_from_data(data), "raw": data}

    return render_template_string("""
    <h2>ç‰©æµè½¨è¿¹æŸ¥è¯¢</h2>
    <form method="post">
      <textarea name="numbers" rows="6" style="width:600px" placeholder="æ¯è¡Œä¸€ä¸ªå•å·ï¼ˆæœ€å¤š30æ¡ï¼‰">{{default_text}}</textarea><br><br>
      <button>æŸ¥è¯¢</button>
    </form>
    {% if message %}<p>{{message}}</p>{% endif %}
    {% if results %}
      <hr><h3>ç»“æœ</h3>
      <div style="white-space:pre-wrap">
        {% for no, r in results.items() %}
          <h4>{{no}}</h4>
          {% if r.tracks %}{{r.tracks}}{% else %}æ— è½¨è¿¹{% endif %}
          <hr>
        {% endfor %}
      </div>
    {% endif %}
    """, results=results, message=message, default_text=default_text)

# ------------------------------
# API
# ------------------------------
@app.route("/api/track/<carrier_id>/<tracking_number>")
def api_track_one(carrier_id, tracking_number):
    data = call_gettrack(carrier_id, tracking_number)
    return app.response_class(json.dumps(data, ensure_ascii=False), mimetype="application/json; charset=utf-8")

@app.route("/api/track_by_agent/<int:agent_id>/<tracking_number>")
def api_track_by_agent(agent_id, tracking_number):
    data = call_gettrack(None, tracking_number, agent_id=agent_id)
    return app.response_class(json.dumps(data, ensure_ascii=False), mimetype="application/json; charset=utf-8")

# ------------------------------
# 17Track API è·¯ç”±
# ------------------------------
@app.route("/api/17track/<tracking_number>")
def api_17track(tracking_number):
    data = call_17track(tracking_number)
    return app.response_class(json.dumps(data, ensure_ascii=False), mimetype="application/json; charset=utf-8")

# ------------------------------
# æ–°å¢è·¯ç”±ï¼Œæ¸²æŸ“å‰ç«¯é¡µé¢
# ------------------------------
@app.route("/frontend")
def frontend_page():
    return render_template("public_frontend.html")

# ------------------------------
# æ–°å¢ JSON APIï¼ˆå‰ç«¯ fetch è°ƒç”¨ç”¨è¿™ä¸ªï¼‰
# ------------------------------
@app.route("/public_track", methods=["POST"])
def public_track_json():
    try:
        data = request.get_json() or {}
        order_id = data.get("order_id", "").strip()

        if not order_id:
            return jsonify({"error": "ç¼ºå°‘ order_id å‚æ•°"}), 400

        s = Shipment.query.filter_by(tracking_number=order_id).first()
        result_data = None
        
        # ä¼˜å…ˆå°è¯•17TrackæŸ¥è¯¢ï¼ˆå¦‚æœæœ‰17Trackå•å·ï¼‰
        if s and s.third_party_tracking_number:
            result_data = call_17track(s.third_party_tracking_number)
        
        # å¦‚æœæ²¡æœ‰17Trackç»“æœï¼Œå°è¯•å¸¸è§„æŸ¥è¯¢
        if not result_data or (isinstance(result_data, dict) and result_data.get("error")):
            if s and s.agent_id:
                result_data = call_gettrack(None, order_id, agent_id=s.agent_id)
            elif s and s.carrier_id:
                result_data = call_gettrack(s.carrier_id, order_id, agent_id=None)
            else:
                result_data = call_gettrack(None, order_id, agent_id=None)

        if isinstance(result_data, dict) and result_data.get("error"):
            return jsonify({"order_id": order_id, "error": result_data.get("error"), "tracks": []})
        else:
            return jsonify({"order_id": order_id, "error": None, "tracks": format_tracks_from_data(result_data)})

    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ------------------------------
# è´¦å•æ‰‹å·¥å¯¼å‡ºï¼ˆä¿ç•™ï¼‰
# ------------------------------
@app.route("/invoice", methods=["GET", "POST"])
@login_required
def invoice():
    if request.method == "GET":
        return render_template_safe("invoice.html")
    customer = request.form.get("customer", "").strip()
    bankinfo = request.form.get("bankinfo", "").strip()
    raw = request.form.get("rows", "").strip()
    if not raw:
        flash("è¯·ç²˜è´´è´¦å•è¡Œ")
        return redirect(url_for("invoice"))
    lines = [l for l in raw.splitlines() if l.strip()]
    parsed = []
    for ln in lines:
        cols = [c.strip() for c in (ln.split("\t") if "\t" in ln else ln.split(","))]
        parsed.append(cols)
    bio = generate_invoice_xlsx(os.getenv("COMPANY_NAME", "å…¬å¸åç§°"),
                                customer or "å®¢æˆ·", bankinfo or "", parsed)
    filename = f"invoice_{int(time.time())}.xlsx"
    return send_file(bio, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ------------------------------
# å¯åŠ¨å‡†å¤‡
# ------------------------------
def ensure_admin_user():
    try:
        if User.query.count() == 0:
            admin_name = os.getenv("ADMIN_USER", "admin")
            admin_pass = os.getenv("ADMIN_PASS", "123456")
            u = User(username=admin_name, is_admin=True)
            u.set_password(admin_pass)
            db.session.add(u)
            db.session.commit()
            app.logger.info(f"å·²åˆ›å»ºé»˜è®¤ç®¡ç†å‘˜è´¦å·ï¼š{admin_name} / {admin_pass}ï¼ˆè¯·éƒ¨ç½²åä¿®æ”¹ï¼‰")
    except Exception:
        app.logger.exception("åˆ›å»ºé»˜è®¤ç®¡ç†å‘˜å¤±è´¥")

# ------------------------------
# ä¸»å…¥å£
# ------------------------------
if __name__ == "__main__":
    with app.app_context():
        db.create_all()
        ensure_sqlite_columns()  # ç§»é™¤ try-exceptï¼Œè®©é”™è¯¯å†’æ³¡ä»¥ä¾¿è°ƒè¯•
        ensure_admin_user()
    app.run(host="0.0.0.0", port=int(os.getenv("PORT", "5000")), debug=True)

flask_app = app
application = app